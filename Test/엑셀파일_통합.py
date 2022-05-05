import openpyxl as xl

# 통합파일 생성
cb_wb = xl.Workbook()
cb_ws = cb_wb.active
cb_ws.title = "Combine"

# 타이틀
cb_ws.append(['순번','제품명','수량','단가','합계'])

# 통합할 파일
tot_file=['11번가','다나와','쿠팡']

# 통합하기
for file in tot_file:
    op_file = f'Test/{file}.xlsx'
    op_wb = xl.load_workbook(op_file,data_only=True)
    op_ws = op_wb.active
    for row in op_ws.iter_rows(min_row=2):
        data = []
        for cell in row:
            data.append(cell.value)
        cb_ws.append(data)

# 순번 업데이트
# for i in range(1, cb_ws.max_row):
#     cb_ws.cell(row=i+1, column=1).value = i   

i = 0
for cell in cb_ws['A']:
    if i != 0:
        cell.value = i
    i=i+1
   

# 파일 저장
cb_wb.save('Test/통합.xlsx')



